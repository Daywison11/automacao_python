#criando planilha
import openpyxl
#criando planilha Book
book  = openpyxl.Workbook()
#visualizando paginas existente
print(book.sheetnames)
#criar pagina
book.create_sheet('cidades')
#selecionar ágina
cidades = book['cidades']

contador = 0
while contador < 20:
    cidades.append(['Ribeirao das neves ', 'mg ' ,' cep', '33820400', '123456'])
    contador +=1
book.save('teste2.xlsx')
