#nao esta usando nesse arquivo
from os import stat
import pandas as pd
import json
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

#esta usando
import time
import random
import requests
import numpy
import Cep
import openpyxl

#CONFIGURAÇÃO PARA CRIA A PLANILHA
#criando planilha Book
book  = openpyxl.Workbook()
#visualizando paginas existente
print(book.sheetnames)
#criar pagina
book.create_sheet('cidades')
#selecionar página
cidades = book['cidades']

#criação da primeira linha
cidades.append(['CIDADE','ESTADO','CEP','IBGE'])
#============================

#=========================
#BUSCA IBGE MAIS CEP
cep = Cep.buscarCeps()
print(cep)
planilha = []
cont = int(0)
arrsize = numpy.size(cep)
#==========================

while cont < arrsize:
    #BUSCA DE DADOS
    #numeros aleatorio
    aleatorio = random.uniform(0,3)
    temp = (round(aleatorio,1))
    #tempo para requests
    time.sleep(temp)
    #url com parametro de pesquisa
    url = "https://viacep.com.br/ws/{}/json/".format(cep[cont])


    #mascara proxi para evitar bloqueio de ip
    proxies = {}
    if cont >= 500  and cont < 900:
        proxies = {'http':'200.229.229.197'}

    elif cont >= 900 and cont < 1200 :
        proxies = {'http':'201.49.83.233'}

    elif cont >= 1200:

        proxies = {'http':'143.208.200.26'}

    else:
        proxies = {'http':'168.232.84.139'}

    print(proxies)
    
    print(url)
    res = requests.get(url, proxies=proxies)
    #retorna o status da aplicação
    status = res.status_code 
    res = res.json()
    erro = res
    
    #verifica se a aplicação esta retornando status 200 senao para o loop
    if status == 200:
    #DECLARANDO DADOS A SER SALVO
        #verifica se a api do ibge retorna erro
        if erro == {'erro': True}:
            #se retornar erro mostra no terminal que nao encontro o ibge
            print('nao encontrado')
            #na tabela (exel) adciona um as colnao como nao encontrado, com o cep pesquisado
            cidades.append(['nao encontrado', 'nao encontrado', cep[cont], 'nao encontrado'])
            cont += 1
        else:
            #caso nao tenha erro pega as variavei do objeto retornado
            ibge = (res['ibge'])
            uf = (res['uf'])
            cepurl = (res['cep'])
            localidade = (res['localidade'])

            #vai adcionar um array com todos os dados de cada cep pesquisados
            planilha.append( localidade + ' ' + uf + ' CEP: ' + cepurl + " IBGE : " +  ibge)

            #print(localidade + s + uf + ' CEP: ' + cepurl + " IBGE : " +  ibge )
            # print no status de busca
            print('buscando IBGE: {}'.format(cont))

            #MONTANDO PLANILHA com os dados
            cidades.append([localidade,uf,cepurl,ibge])#cada  dado e uma coluna
            cont += 1
    else:
        break


#print no arr com todas as planilhas
print(planilha)

#SALVANDO PLANILHA com nome e extensao
book.save('teste20.xlsx')